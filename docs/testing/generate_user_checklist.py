"""Generate simple user-facing MAX RASS checklist."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).with_name("max-rass-user-checklist.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="E8F5E9")
TITLE_FONT = Font(bold=True, size=16, color="2E7D32")
HINT_FONT = Font(italic=True, color="666666", size=10)
THIN = Side(style="thin", color="C8E6C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

ITEMS = [
    ("Вход в систему", "Открывается страница входа", "Перейти на http://localhost:5173/"),
    ("Вход в систему", "Вход по логину и паролю", "Ввести логин и пароль → нажать «Войти»"),
    ("Вход в систему", "Ошибка при неверном пароле", "Ввести неправильный пароль → появляется сообщение об ошибке"),
    ("Вход в систему", "Выход из системы", "Нажать «Выйти» → снова видна страница входа"),
    ("Вход в систему", "Вход через MAX", "Нажать «Войти через MAX» → открывается бот, вход подтверждается"),
    ("Вход в систему", "Восстановление пароля", "«Забыли пароль?» → подтвердить в MAX → войти с новым паролем"),
    ("HR-панель", "Открывается после входа HR", "Войти как hr.manager / hr123456"),
    ("HR-панель", "Список учеников", "Вкладка «Ученики» — видны ученики, можно искать и фильтровать"),
    ("HR-панель", "Добавить ученика", "Кнопка добавления → заполнить форму → ученик появился в списке"),
    ("HR-панель", "Рабочие группы", "Вкладка «Группы» — список групп, можно открыть состав"),
    ("HR-панель", "Добавить группу", "Создать новую группу → она отображается в списке"),
    ("HR-панель", "Добавить ученика в группу", "Открыть группу → добавить участника"),
    ("HR-панель", "Расписание занятий", "Вкладка «Расписание» — видны занятия"),
    ("HR-панель", "Страйки", "Вкладка «Страйки» — видны страйки, можно выписать или отменить"),
    ("HR-панель", "Отчёты", "Вкладка «Отчёты» — открываются без ошибок, данные понятны"),
    ("Панель преподавателя", "Открывается после входа", "Войти как teacher.demo / teacher123456"),
    ("Панель преподавателя", "Ближайшее занятие на главной", "На главной видно следующее занятие или пустое состояние"),
    ("Панель преподавателя", "Расписание", "Вкладка «Расписание» — занятия отображаются по дням"),
    ("Панель преподавателя", "Создать занятие", "«Назначить занятие» → заполнить → занятие появилось"),
    ("Панель преподавателя", "Отметить посещаемость", "Выбрать занятие → отметить учеников → «Сохранить»"),
    ("Панель преподавателя", "История посещаемости", "Кнопка «История» — показывает прошлые отметки"),
    ("Панель преподавателя", "Отчёт", "Вкладка «Отчёт» — данные по группам/ученикам"),
    ("MAX-бот", "Приветствие", "Написать боту — приходит ответ с инструкцией"),
    ("MAX-бот", "Подтверждение входа", "При входе через MAX — бот просит подтвердить"),
    ("MAX-бот", "Напоминание о занятии", "За сутки или за 3 часа приходит напоминание"),
]


def style_row(ws, row: int, cols: int, fill=None) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = BORDER
        cell.alignment = WRAP
        if fill:
            cell.fill = fill


def build_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Чеклист"

    ws.merge_cells("A1:D1")
    ws["A1"] = "MAX RASS — чеклист для пользователя"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(vertical="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        "Отметьте «Да», если всё работает как описано. "
        "Сайт: http://localhost:5173/  ·  "
        "HR: hr.manager / hr123456  ·  Преподаватель: teacher.demo / teacher123456"
    )
    ws["A2"].font = HINT_FONT
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 36

    headers = ["№", "Раздел", "Что проверить", "Как проверить", "Готово", "Замечания"]
    ws.append([])
    ws.append(headers)
    header_row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    current_section = None
    for idx, (section, what, how) in enumerate(ITEMS, start=1):
        row = header_row + idx
        ws.append([idx, section, what, how, "", ""])
        if section != current_section:
            current_section = section
            style_row(ws, row, len(headers), SECTION_FILL)
        else:
            style_row(ws, row, len(headers))
        ws.cell(row=row, column=1).alignment = CENTER

    widths = {1: 5, 2: 22, 3: 30, 4: 42, 5: 10, 6: 28}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"

    dv = DataValidation(type="list", formula1='"Да,Нет,—"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"E5:E{ws.max_row}")

    ws.row_dimensions[1].height = 28


def main() -> None:
    wb = Workbook()
    build_sheet(wb)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
