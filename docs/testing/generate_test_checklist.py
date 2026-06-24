"""Generate MAX RASS testing checklist Excel workbook."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT = Path(__file__).with_name("max-rass-test-checklist.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(bold=True, size=11, color="1F4E79")
THIN = Side(style="thin", color="B4C6E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def set_col_widths(ws, widths: dict[int, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_instruction_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Инструкция"

    rows = [
        ("MAX RASS — мини-инструкция для тестирования", ""),
        ("", ""),
        ("1. Подготовка окружения", ""),
        ("Требования", "Docker Desktop, PowerShell, браузер Chrome/Edge"),
        ("Клонирование / каталог", "cd app"),
        ("Конфигурация", "Скопируйте .env.example → .env, при необходимости заполните MAX_BOT_TOKEN и MAX_USER_MAP_JSON"),
        ("Сборка фронтенда", "cd web && npm install && npm run build"),
        ("Запуск стенда", "cd app && docker compose up -d --build"),
        ("Полный сброс (БД + Keycloak)", "docker compose down -v && docker compose up -d --build"),
        ("", ""),
        ("2. Адреса сервисов", ""),
        ("Приложение (вход)", "http://localhost:5173/"),
        ("HR тест-панель", "http://localhost:5173/hr-test.html"),
        ("Преподаватель тест-панель", "http://localhost:5173/teacher-test.html"),
        ("Тест БД", "http://localhost:5173/db-test"),
        ("Keycloak Admin", "http://localhost:5173/admin/ (admin / admin)"),
        ("MAX Auth API health", "http://localhost:8025/health"),
        ("App PostgreSQL", "localhost:5433, БД max_rass, user/pass: max_rass"),
        ("", ""),
        ("3. Тестовые учётные записи (после импорта realm)", ""),
        ("hr.manager / hr123456", "Роль HR — панель /hr-test.html"),
        ("teacher.demo / teacher123456", "Роль Преподаватель — /teacher-test.html"),
        ("admin / admin123456", "Роль Администратор"),
        ("", ""),
        ("4. Быстрый сценарий smoke-теста", ""),
        ("Шаг 1", "Проверить health всех сервисов (лист «Чеклист», раздел 1)"),
        ("Шаг 2", "GET /api/v1/db/health → {\"status\":\"ok\"}"),
        ("Шаг 3", "На /db-test нажать «Заполнить demo»"),
        ("Шаг 4", "Войти как hr.manager → открыть /hr-test.html → «Обновить всё»"),
        ("Шаг 5", "Войти как teacher.demo → /teacher-test.html → загрузить группы и расписание"),
        ("Шаг 6", "Проверить вход через MAX (если настроен MAX_BOT_TOKEN и MAX_USER_MAP_JSON)"),
        ("", ""),
        ("5. Полезные команды", ""),
        ("Логи max-auth", "docker compose logs -f max-auth"),
        ("psql в app-postgres", "docker compose exec app-postgres psql -U max_rass -d max_rass"),
        ("Ручной запуск напоминаний", "POST http://localhost:8025/api/v1/db/lesson-reminders/run"),
        ("Проверка таблиц", "curl http://localhost:5173/api/v1/db/tables"),
        ("", ""),
        ("6. Документация", ""),
        ("README стенда", "app/README.md"),
        ("Мануал тестирования БД", "docs/testing/database-test-manual.md"),
        ("Схема БД", "docs/architecture/max-rass-database-schema.html"),
    ]

    for i, (a, b) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        if a and a[0].isdigit() and a[1] == ".":
            ws.cell(row=i, column=1).font = SUBTITLE_FONT

    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    set_col_widths(ws, {1: 34, 2: 72})
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER


CHECKLIST = [
    # (section, test_case, steps, expected)
    ("1. Инфраструктура", "Docker Compose поднял все сервисы", "docker compose ps — все сервисы healthy/running", "app-postgres, postgres, keycloak, max-auth, web — Up"),
    ("1. Инфраструктура", "Health MAX Auth API", "GET http://localhost:8025/health", "HTTP 200, status ok"),
    ("1. Инфраструктура", "Web доступен", "Открыть http://localhost:5173/", "Страница входа загружается без ошибок"),
    ("1. Инфраструктура", "Keycloak Admin доступен", "http://localhost:5173/admin/ → admin/admin", "Консоль Keycloak открывается, realm max-education"),
    ("1. Инфраструктура", "Nginx проксирует API", "GET http://localhost:5173/api/v1/db/health", "JSON {\"status\":\"ok\"}"),
    ("2. База данных", "Схема app применена", "GET /api/v1/db/tables", "Таблицы: users, profiles, roles, groups, lessons, attendance_marks, strikes, notifications и др."),
    ("2. База данных", "Seed demo-данных", "POST /api/v1/db/seed-demo или кнопка на /db-test", "Успех, 4 профиля, группа Backend-26, занятие, страйк"),
    ("2. База данных", "Read-only SQL запрос", "На /db-test выполнить SELECT из app.profiles LIMIT 5", "Результат ≤200 строк, без ошибок"),
    ("2. База данных", "Запрет изменяющих SQL", "Попытка INSERT/DELETE на /db-test", "Ошибка: запрещённый тип запроса"),
    ("2. База данных", "Профили demo после seed", "SELECT p.last_name, r.code FROM app.profiles p JOIN app.roles r ON r.id=p.role_id", "4 записи; у Ким Мария phone=NULL, id_curator заполнен"),
    ("2. База данных", "Посещаемость demo", "Проверить attendance_marks для demo-занятия", "Статус late для участника Ким"),
    ("2. База данных", "Уведомление demo", "SELECT kind FROM app.notifications", "Есть lesson_reminder_1d для куратора"),
    ("3. Авторизация Keycloak", "Вход по логину/паролю HR", "http://localhost:5173/ → hr.manager / hr123456", "Редирект на HR dashboard или успешная авторизация"),
    ("3. Авторизация Keycloak", "Вход преподавателя", "teacher.demo / teacher123456", "Доступ к teacher-test.html"),
    ("3. Авторизация Keycloak", "Неверный пароль", "Ввести неверный пароль", "Сообщение об ошибке, вход не выполнен"),
    ("3. Авторизация Keycloak", "GET /api/v1/auth/me", "После входа — запрос с Bearer token", "Возвращает username и роли пользователя"),
    ("3. Авторизация Keycloak", "Выход", "Нажать «Выйти» на тест-панели", "Сессия завершена, редирект на /auth.html"),
    ("3. Авторизация Keycloak", "Смена временного пароля", "Создать user с Temporary password в Keycloak, первый вход", "Keycloak предлагает сменить пароль"),
    ("4. Вход через MAX", "Старт MAX-сессии", "POST /api/v1/auth/max/start или кнопка «Войти через MAX»", "session_id, bot_url, expires_at"),
    ("4. Вход через MAX", "Polling статуса", "GET /api/v1/auth/max/status/{session_id}", "status: pending → confirmed после подтверждения"),
    ("4. Вход через MAX", "Подтверждение в боте", "Открыть bot_url, подтвердить вход (или демо-кнопка)", "status=confirmed, keycloak_username заполнен"),
    ("4. Вход через MAX", "Exchange токена", "POST /api/v1/auth/max/exchange после confirm", "Access token Keycloak, редирект в приложение"),
    ("4. Вход через MAX", "Неизвестный MAX ID", "Confirm с ID не из MAX_USER_MAP_JSON", "Отказ / сообщение неизвестному пользователю"),
    ("4. Вход через MAX", "Истечение сессии", "Не подтверждать 5+ минут", "status=expired"),
    ("5. Сброс пароля MAX", "Старт сброса", "Ссылка «Забыли пароль?» → /max?mode=reset", "Сессия reset, ссылка на бота"),
    ("5. Сброс пароля MAX", "Подтверждение сброса в боте", "Подтвердить сброс в MAX", "status=password_reset, temp_password в ответе"),
    ("5. Сброс пароля MAX", "Вход с временным паролем", "Войти с temp_password на сайте", "Успешный вход, required action update_password"),
    ("5. Сброс пароля MAX", "Смена пароля", "POST /api/v1/auth/change-password", "Пароль обновлён, повторный вход с новым паролем"),
    ("6. HR API", "GET /api/v1/hr/me", "Войти как HR, запрос /me", "Профиль HR-пользователя"),
    ("6. HR API", "GET /api/v1/hr/users", "Список пользователей на /hr-test.html", "Таблица пользователей с ролями"),
    ("6. HR API", "POST /api/v1/hr/users", "Создать пользователя (фамилия, имя, роль, телефон)", "201, пользователь в списке"),
    ("6. HR API", "PATCH /api/v1/hr/users/{id}", "Изменить телефон/куратора", "Данные обновлены"),
    ("6. HR API", "GET /api/v1/hr/groups", "Вкладка «Группы»", "Список групп включая Backend-26"),
    ("6. HR API", "POST /api/v1/hr/groups", "Создать новую группу", "Группа появилась в списке"),
    ("6. HR API", "POST /api/v1/hr/groups/{id}/members", "Добавить участника в группу", "Участник привязан"),
    ("6. HR API", "DELETE member из группы", "Удалить участника", "Участник удалён"),
    ("6. HR API", "GET /api/v1/hr/strikes", "Вкладка «Страйки»", "Demo-страйк late отображается"),
    ("6. HR API", "POST /api/v1/hr/strikes", "Выписать новый страйк", "Страйк создан со статусом active"),
    ("6. HR API", "POST strikes/revoke", "Отменить страйк", "status=revoked"),
    ("6. HR API", "Апелляции", "POST appeal → resolve на /hr-test", "Статус апелляции обновлён"),
    ("6. HR API", "GET /api/v1/hr/reports/summary", "Вкладка «Отчёты» → сводка", "Числовые показатели без ошибок"),
    ("6. HR API", "Отчёты посещаемости", "GET reports/attendance/groups и /users", "Данные по группам и сотрудникам"),
    ("6. HR API", "Доступ без роли HR", "Запрос HR API под teacher.demo", "HTTP 403 Forbidden"),
    ("7. Teacher API", "GET /api/v1/teacher/me", "Войти как teacher.demo", "Профиль преподавателя"),
    ("7. Teacher API", "GET /api/v1/teacher/groups", "Вкладка «Группы»", "Рабочие группы преподавателя"),
    ("7. Teacher API", "GET /api/v1/teacher/lessons", "Вкладка «Расписание»", "Список занятий"),
    ("7. Teacher API", "POST /api/v1/teacher/lessons", "Создать занятие (группа, время, аудитория)", "Занятие в расписании"),
    ("7. Teacher API", "PATCH /api/v1/teacher/lessons/{id}", "Изменить время/аудиторию", "Данные обновлены"),
    ("7. Teacher API", "DELETE /api/v1/teacher/lessons/{id}", "Удалить занятие", "Занятие удалено"),
    ("7. Teacher API", "GET attendance", "GET /lessons/{id}/attendance", "Список участников и статусов"),
    ("7. Teacher API", "POST attendance", "Отметить present/late/absent", "Отметки сохранены в БД"),
    ("7. Teacher API", "История посещаемости", "GET /attendance/history", "История по фильтрам"),
    ("7. Teacher API", "Отчёты", "GET reports/attendance/groups и /users", "Отчёты без ошибок"),
    ("7. Teacher API", "Доступ без роли teacher", "HR API под hr.manager на teacher endpoints", "HTTP 403"),
    ("8. MAX бот", "Webhook events", "POST /api/v1/bot/events (если настроен токен)", "200 OK, событие обработано"),
    ("8. MAX бот", "Приветствие нового пользователя", "Первое сообщение неизвестному MAX ID", "Контакт HR / инструкция"),
    ("8. MAX бот", "Кнопки сотрудника", "Callback-кнопки в боте (расписание и т.д.)", "Корректный ответ бота"),
    ("8. MAX бот", "Подтверждение входа login_", "Payload login_{session}", "Сессия confirmed"),
    ("8. MAX бот", "Подтверждение сброса reset_", "Payload reset_{session}", "temp_password отправлен"),
    ("9. Напоминания о занятиях", "Worker включён", "LESSON_REMINDER_ENABLED=true в .env", "В логах max-auth: worker started"),
    ("9. Напоминания о занятиях", "Ручной прогон", "POST /api/v1/db/lesson-reminders/run", "Ответ с количеством отправленных"),
    ("9. Напоминания о занятиях", "Напоминание за 24ч", "Создать занятие через ~24ч, дождаться окна ±2 мин", "notification kind=lesson_reminder_1d, сообщение в MAX"),
    ("9. Напоминания о занятиях", "Напоминание за 3ч", "Занятие через ~3ч", "kind=lesson_reminder_3h"),
    ("9. Напоминания о занятиях", "Без телефона → куратору", "Участник без phone (demo Ким)", "Уведомление delivered_to = id_curator"),
    ("9. Напоминания о занятиях", "Нет дублей", "Повторный run в том же окне", "Повторное уведомление не создаётся"),
]


def build_checklist_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Чеклист")
    headers = ["№", "Раздел", "Тест-кейс", "Шаги", "Ожидаемый результат", "Статус", "Комментарий", "Тестировщик", "Дата"]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    current_section = None
    for idx, (section, case, steps, expected) in enumerate(CHECKLIST, start=1):
        ws.append([idx, section, case, steps, expected, "", "", "", ""])
        row = idx + 1
        if section != current_section:
            current_section = section
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = SECTION_FILL

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER
        ws.cell(row=cell.row, column=1).alignment = CENTER

    set_col_widths(ws, {1: 5, 2: 22, 3: 28, 4: 38, 5: 32, 6: 12, 7: 24, 8: 16, 9: 12})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,Skip,N/A"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Выберите Pass, Fail, Skip или N/A"
    dv.errorTitle = "Неверный статус"
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")


def build_summary_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Сводка")
    ws["A1"] = "Сводка тестирования MAX RASS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    labels = [
        ("Версия / сборка:", ""),
        ("Дата тестирования:", ""),
        ("Тестировщик:", ""),
        ("Окружение:", "localhost (Docker)"),
        ("", ""),
        ("Всего тест-кейсов:", f"=COUNTA(Чеклист!A2:A{len(CHECKLIST)+1})"),
        ("Pass:", f'=COUNTIF(Чеклист!F2:F{len(CHECKLIST)+1},"Pass")'),
        ("Fail:", f'=COUNTIF(Чеклист!F2:F{len(CHECKLIST)+1},"Fail")'),
        ("Skip:", f'=COUNTIF(Чеклист!F2:F{len(CHECKLIST)+1},"Skip")'),
        ("N/A:", f'=COUNTIF(Чеклист!F2:F{len(CHECKLIST)+1},"N/A")'),
        ("Не проверено:", f'=COUNTIF(Чеклист!F2:F{len(CHECKLIST)+1},"")'),
        ("", ""),
        ("Примечания:", ""),
    ]
    for i, (label, value) in enumerate(labels, start=3):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
        if label.endswith(":") and not value.startswith("="):
            ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor="FFF2CC")

    set_col_widths(ws, {1: 22, 2: 40, 3: 20})
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = WRAP
            cell.border = BORDER


def main() -> None:
    wb = Workbook()
    build_instruction_sheet(wb)
    build_checklist_sheet(wb)
    build_summary_sheet(wb)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    main()
